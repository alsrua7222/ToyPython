import parserBOJ
import openpyxl as xl

parse = parserBOJ.Parse()
wb = xl.Workbook()
ws = wb.active
ws.title = "sheet100"

# 컬럼명 지정
col_names = ['문제 번호', '문제 제목', '맞힌 사람', '제출 횟수', '정답률']
for seq, name in enumerate(col_names):
    ws.cell(row=1, column=seq+1, value=name)

row_num = 2
# 데이터 입력
for n, rows in enumerate(parse.processParsing()):
    for seq, value in enumerate(rows):
        ws.cell(row=row_num+n, column=seq+1, value=value)
wb.save("BOJ_Parsing.xlsx")
wb.close()